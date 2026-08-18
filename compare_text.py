from openpyxl import load_workbook

# Template
template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws_t = template.active

# Generated
generated = load_workbook('generated_fsr/FSR_Devanadera_20260817_145821.xlsx')
ws_g = generated.active

print("\n=== TEXT COMPARISON (Template row 27-41 vs Generated row 20-34) ===\n")

row_map = [
    (27, 20, "A", "Concurrent teaching header"),
    (30, 23, "A", "(NONE) - College outside UP"),
    (30, 23, "E", "(NONE) - No. of subjects"),
    (30, 23, "H", "(NONE) - No. of units"),
    (31, 24, "A", "COLLEGE OUTSIDE U.P. SYSTEM"),
    (31, 24, "E", "No. of subjects label"),
    (31, 24, "H", "No. of units label"),
    (33, 26, "A", "(NONE) - UP College"),
    (34, 27, "A", "U.P. COLLEGE/DEPT."),
    (35, 28, "A", "NOTE text"),
    (36, 29, "H", "Certified Correct:"),
    (39, 32, "G", "Registrar name"),
    (40, 33, "G", "University Registrar"),
]

for t_row, g_row, col, desc in row_map:
    t_val = ws_t[f'{col}{t_row}'].value
    g_val = ws_g[f'{col}{g_row}'].value
    
    # Check formatting too
    t_bold = ws_t[f'{col}{t_row}'].font.bold
    g_bold = ws_g[f'{col}{g_row}'].font.bold
    
    match = "✓" if str(t_val).strip() == str(g_val).strip() else "✗"
    bold_match = "✓" if t_bold == g_bold else "✗"
    
    print(f"{match} {desc}")
    print(f"  Template: '{t_val}' (bold={t_bold})")
    print(f"  Generated: '{g_val}' (bold={g_bold})")
    if match == "✗" or bold_match == "✗":
        print(f"  ⚠️ MISMATCH!")
    print()
