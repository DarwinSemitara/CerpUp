from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load template
wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("=" * 120)
print("FSRFORMAT.XLSX STRUCTURE SCAN: ROWS 50-69")
print("=" * 120)

# Scan rows 50-69
for row in range(50, 70):
    print(f"\n{'=' * 120}")
    print(f"ROW {row}")
    print("=" * 120)
    
    # Show merged cells in this row
    merged_in_row = []
    for merge in ws.merged_cells.ranges:
        if merge.min_row == row:
            merged_in_row.append(str(merge))
    
    if merged_in_row:
        print(f"  MERGES: {', '.join(merged_in_row)}")
    else:
        print(f"  MERGES: None")
    
    # Show content and borders for each column A-K
    print(f"\n  COLUMNS:")
    for col_idx in range(1, 12):  # A to K
        col = get_column_letter(col_idx)
        cell = ws[f'{col}{row}']
        
        # Value
        value = str(cell.value)[:50] if cell.value else "(empty)"
        
        # Check if part of merge
        is_merged = "NO"
        merge_range = ""
        for merge in ws.merged_cells.ranges:
            if cell.coordinate in merge:
                is_merged = "YES"
                merge_range = f" [{merge}]"
                break
        
        # Border info
        border_parts = []
        if cell.border:
            if cell.border.left and cell.border.left.style:
                border_parts.append(f"L:{cell.border.left.style}")
            if cell.border.right and cell.border.right.style:
                border_parts.append(f"R:{cell.border.right.style}")
            if cell.border.top and cell.border.top.style:
                border_parts.append(f"T:{cell.border.top.style}")
            if cell.border.bottom and cell.border.bottom.style:
                border_parts.append(f"B:{cell.border.bottom.style}")
        
        border_str = ", ".join(border_parts) if border_parts else "none"
        
        print(f"    {col}{row}: Merged={is_merged:3}{merge_range:15} | Borders: {border_str:25} | Value: {value}")

print("\n" + "=" * 120)
print("SUMMARY")
print("=" * 120)

# Identify key rows
print("\nKey row identification:")
for row in range(50, 70):
    cell_a = ws[f'A{row}'].value
    if cell_a:
        cell_str = str(cell_a)[:80]
        if 'IMPLEMENTATION' in cell_str.upper():
            print(f"  Row {row}: SECTION HEADER - {cell_str}")
        elif 'TITLE' in cell_str.upper() and 'COMPLETE' in cell_str.upper():
            print(f"  Row {row}: COLUMN HEADER - {cell_str}")
        elif 'Total Research' in cell_str:
            print(f"  Row {row}: TOTAL ROW - {cell_str}")
        elif any(keyword in cell_str for keyword in ['Proposal', 'SAMPLE', 'Lorem', 'Study']):
            print(f"  Row {row}: DATA ROW - {cell_str}")
