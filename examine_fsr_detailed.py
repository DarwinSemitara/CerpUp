import openpyxl
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook('static/reference/SAMPLE FSR.xlsx')
ws = wb.active

print("="*100)
print("DETAILED FSR STRUCTURE ANALYSIS")
print("="*100)

# Function to check if row has significant content


def has_content(row_idx, ws):
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and str(cell.value).strip():
            return True
    return False


# Track section headers
sections = []
current_section = None

for row_idx in range(1, min(200, ws.max_row + 1)):
    if not has_content(row_idx, ws):
        continue

    # Get cell A (first column)
    cell_a = ws.cell(row=row_idx, column=1)
    value_a = str(cell_a.value) if cell_a.value else ""

    # Check if it's a section header (starts with Roman numeral or specific pattern)
    if value_a.startswith(('I.', 'II.', 'III.', 'IV.', 'V.', 'VI.')) or 'SECTION' in value_a.upper():
        sections.append({
            'row': row_idx,
            'title': value_a,
            'bold': cell_a.font.bold if cell_a.font else False
        })
        current_section = value_a
        print(f"\n{'='*100}")
        print(f"ROW {row_idx}: {value_a}")
        print(f"{'='*100}")

    # Print row details
    row_data = []
    for col_idx in range(1, 13):  # First 12 columns
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            col_letter = get_column_letter(col_idx)
            formatting = []
            if cell.font and cell.font.bold:
                formatting.append('B')
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != '00000000':
                formatting.append('F')
            if cell.border and (cell.border.left.style or cell.border.top.style):
                formatting.append('BR')

            format_str = f"[{','.join(formatting)}]" if formatting else ""
            value_str = str(cell.value)[:60]
            row_data.append(f"{col_letter}:{value_str}{format_str}")

    if row_data:
        print(f"  Row {row_idx:3d}: {' | '.join(row_data)}")

print("\n" + "="*100)
print("SECTION SUMMARY")
print("="*100)
for idx, section in enumerate(sections, 1):
    print(f"{idx}. Row {section['row']:4d}: {section['title']}")

print("\n" + "="*100)
print("KEY PATTERNS IDENTIFIED")
print("="*100)

# Identify table headers (rows with multiple bordered cells)
table_headers = []
for row_idx in range(1, min(200, ws.max_row + 1)):
    bordered_count = 0
    for col_idx in range(1, 13):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.border and (cell.border.top.style or cell.border.bottom.style):
            if cell.value:
                bordered_count += 1

    if bordered_count >= 4:  # If 4+ cells are bordered, it's likely a table header
        first_cell = ws.cell(row=row_idx, column=1)
        if first_cell.value:
            table_headers.append({
                'row': row_idx,
                'value': str(first_cell.value)[:50]
            })

print("\nTABLE HEADERS (rows with multiple bordered cells):")
for header in table_headers[:20]:  # First 20
    print(f"  Row {header['row']:3d}: {header['value']}")
