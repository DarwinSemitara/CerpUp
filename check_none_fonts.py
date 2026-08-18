from openpyxl import load_workbook

template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws_t = template.active

generated = load_workbook('generated_fsr/FSR_Devanadera_20260817_145821.xlsx')
ws_g = generated.active

print("\n=== FONT CHECK FOR (NONE) AND LABELS ===\n")

cells_to_check = [
    ('A30', 'A23', "(NONE) left"),
    ('E30', 'E23', "(NONE) middle"),
    ('H30', 'H23', "(NONE) right"),
    ('E31', 'E24', "No. of subjects (row 31)"),
    ('H31', 'H24', "No. of units (row 31)"),
    ('A33', 'A26', "(NONE) left row 33"),
    ('E33', 'E26', "(NONE) middle row 33"),
    ('H33', 'H26', "(NONE) right row 33"),
    ('E34', 'E27', "No. of subjects (row 34)"),
    ('H34', 'H27', "No. of units (row 34)"),
]

for t_cell, g_cell, desc in cells_to_check:
    t = ws_t[t_cell]
    g = ws_g[g_cell]
    
    print(f"{desc}:")
    print(f"  Template {t_cell}: font={t.font.name}, bold={t.font.bold}, value='{t.value}'")
    print(f"  Generated {g_cell}: font={g.font.name}, bold={g.font.bold}, value='{g.value}'")
    
    if t.font.name != g.font.name or t.font.bold != g.font.bold:
        print(f"  ⚠️ MISMATCH! Should be: {t.font.name}, bold={t.font.bold}")
    print()
