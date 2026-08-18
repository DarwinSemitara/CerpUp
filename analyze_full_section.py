from openpyxl import load_workbook

wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("\n=== FULL CONCURRENT TEACHING SECTION (rows 27-41) ===\n")

for row_num in range(27, 42):
    print(f"ROW {row_num}:")
    
    # Check each cell A-K
    for col_num in range(1, 12):  # A=1 to K=11
        cell = ws.cell(row_num, col_num)
        col_letter = chr(64 + col_num)  # A=65
        
        if isinstance(cell, type(cell)) and hasattr(cell, 'value'):
            if cell.value is not None and cell.value not in ['None', '']:
                # Check if part of a merge
                is_merged = False
                merge_range = None
                for merge in ws.merged_cells.ranges:
                    if cell.coordinate in merge:
                        is_merged = True
                        merge_range = str(merge)
                        break
                
                if is_merged:
                    print(f"  {col_letter}{row_num}: '{cell.value}' (MERGED: {merge_range})")
                else:
                    print(f"  {col_letter}{row_num}: '{cell.value}'")
    
    print()
