from openpyxl import load_workbook

# Load template
template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws_template = template.active

# Load generated FSR
generated = load_workbook('generated_fsr/FSR_Devanadera_20260817_135352.xlsx')
ws_generated = generated.active

print('=== TEMPLATE (FSRFORMAT.xlsx) ===')
print('Rows around footnotes and concurrent teaching (19-26):')
for row in range(19, 27):
    cell = ws_template.cell(row=row, column=1)
    print(f'Row {row}: "{cell.value}"')

print('\n=== GENERATED FSR ===')
print('Same section after row deletions (16-23):')
for row in range(16, 24):
    cell = ws_generated.cell(row=row, column=1)
    print(f'Row {row}: "{cell.value}"')
    
print('\n=== MERGED CELLS COMPARISON ===')
print('Template merged cells (rows 19-26):')
for merged_range in ws_template.merged_cells.ranges:
    if 19 <= merged_range.min_row <= 26:
        print(f'  {merged_range}')

print('\nGenerated merged cells (rows 16-23):')
for merged_range in ws_generated.merged_cells.ranges:
    if 16 <= merged_range.min_row <= 23:
        print(f'  {merged_range}')
