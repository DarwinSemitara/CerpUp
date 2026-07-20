import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json

# Load the workbook
wb = openpyxl.load_workbook('static/reference/SAMPLE FSR.xlsx')

print("=" * 80)
print("EXCEL FILE ANALYSIS - SAMPLE FSR.xlsx")
print("=" * 80)

# Sheet information
print(f"\n📄 SHEETS: {wb.sheetnames}")
print(f"   Active sheet: {wb.active.title}")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
    
    # Check for merged cells
    if ws.merged_cells:
        print(f"\n🔗 MERGED CELLS ({len(ws.merged_cells.ranges)}):")
        for merged_range in list(ws.merged_cells.ranges)[:20]:  # Show first 20
            print(f"   {merged_range}")
    
    # Analyze content and formatting
    print(f"\n📋 CONTENT & FORMATTING (First 50 rows):\n")
    
    for row_idx in range(1, min(51, ws.max_row + 1)):
        row_data = []
        has_content = False
        
        for col_idx in range(1, min(20, ws.max_column + 1)):  # First 20 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.value is not None:
                has_content = True
                
                # Get formatting info
                cell_info = {
                    'value': str(cell.value),
                    'col': get_column_letter(col_idx),
                }
                
                # Font info
                if cell.font:
                    if cell.font.bold:
                        cell_info['bold'] = True
                    if cell.font.size and cell.font.size != 11:
                        cell_info['size'] = cell.font.size
                    if cell.font.color and cell.font.color.rgb:
                        cell_info['color'] = str(cell.font.color.rgb)
                
                # Fill color
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                    cell_info['fill'] = str(cell.fill.start_color.rgb)
                
                # Alignment
                if cell.alignment:
                    if cell.alignment.horizontal:
                        cell_info['align'] = cell.alignment.horizontal
                    if cell.alignment.wrap_text:
                        cell_info['wrap'] = True
                
                # Border
                if cell.border and (cell.border.left.style or cell.border.right.style or 
                                  cell.border.top.style or cell.border.bottom.style):
                    cell_info['border'] = True
                
                row_data.append(cell_info)
        
        if has_content:
            print(f"\nRow {row_idx}:")
            for cell_info in row_data:
                formatting = []
                if cell_info.get('bold'):
                    formatting.append('BOLD')
                if cell_info.get('size'):
                    formatting.append(f"Size:{cell_info['size']}")
                if cell_info.get('fill'):
                    formatting.append(f"Fill:{cell_info['fill']}")
                if cell_info.get('align'):
                    formatting.append(f"Align:{cell_info['align']}")
                if cell_info.get('wrap'):
                    formatting.append('WRAP')
                if cell_info.get('border'):
                    formatting.append('BORDER')
                
                format_str = f" [{', '.join(formatting)}]" if formatting else ""
                print(f"  {cell_info['col']}: {cell_info['value'][:100]}{format_str}")
    
    # Column widths
    print(f"\n📏 COLUMN WIDTHS:")
    for col_idx in range(1, min(20, ws.max_column + 1)):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        if width:
            print(f"   {col_letter}: {width}")
    
    # Row heights
    print(f"\n📐 ROW HEIGHTS (non-default):")
    for row_idx in range(1, min(51, ws.max_row + 1)):
        height = ws.row_dimensions[row_idx].height
        if height and height != 15:  # Default is usually 15
            print(f"   Row {row_idx}: {height}")

print("\n" + "="*80)
print("ANALYSIS COMPLETE")
print("="*80)
