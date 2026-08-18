from openpyxl import load_workbook

wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("\n=== CONCURRENT TEACHING SECTION (rows 27-41, cols A-K) ===\n")

for row_num in range(27, 42):
    print(f"ROW {row_num}:")
    
    for col_num in range(1, 12):  # A=1 to K=11
        cell = ws.cell(row_num, col_num)
        col_letter = chr(64 + col_num)
        
        # Check if merged
        is_merged = False
        merge_info = None
        for merge in ws.merged_cells.ranges:
            if cell.coordinate in merge:
                is_merged = True
                # Check if this is the top-left cell
                if merge.min_row == row_num and merge.min_col == col_num:
                    merge_info = f"TOP-LEFT of {merge}"
                else:
                    merge_info = f"part of {merge}"
                break
        
        if is_merged:
            top_left = ws.cell(merge.min_row, merge.min_col)
            print(f"  {col_letter}{row_num}: MERGED ({merge_info}) - value in top-left: '{top_left.value}'")
        else:
            print(f"  {col_letter}{row_num}: '{cell.value}'")
    
    print()

print("\n=== MERGED CELL RANGES (rows 27-41) ===\n")
for merge in ws.merged_cells.ranges:
    if 27 <= merge.min_row <= 41:
        top_left = ws.cell(merge.min_row, merge.min_col)
        print(f"{merge}: '{top_left.value}'")
