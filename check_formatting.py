from openpyxl import load_workbook

# Check template formatting
template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws_template = template.active

# Check generated formatting
generated = load_workbook('generated_fsr/FSR_Devanadera_20260817_145821.xlsx')
ws_gen = generated.active

print("\n=== TEMPLATE ROW 27 (A27) FORMATTING ===")
cell_t = ws_template['A27']
print(
    f"Font: {cell_t.font.name}, Size: {cell_t.font.size}, Bold: {cell_t.font.bold}")
print(
    f"Alignment: horizontal={cell_t.alignment.horizontal}, vertical={cell_t.alignment.vertical}")
print(f"Fill: {cell_t.fill.patternType}, {cell_t.fill.fgColor.rgb if cell_t.fill.fgColor else 'None'}")
print(f"Border: {cell_t.border}")

print("\n=== GENERATED ROW 20 (A20) FORMATTING ===")
cell_g = ws_gen['A20']
print(
    f"Font: {cell_g.font.name}, Size: {cell_g.font.size}, Bold: {cell_g.font.bold}")
print(
    f"Alignment: horizontal={cell_g.alignment.horizontal}, vertical={cell_g.alignment.vertical}")
print(f"Fill: {cell_g.fill.patternType}, {cell_g.fill.fgColor.rgb if cell_g.fill.fgColor else 'None'}")
print(f"Border: {cell_g.border}")
