from openpyxl import load_workbook

template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = template.active

print("\n=== RESEARCH PROPOSAL SECTION STRUCTURE (rows 42-49) ===\n")

for row_num in range(42, 50):
    print(f"\n{'='*80}")
    print(f"ROW {row_num}:")
    print(f"{'='*80}")
    
    for col_num in range(1, 12):  # A=1 to K=11
        cell = ws.cell(row_num, col_num)
        col_letter = chr(64 + col_num)
        
        # Check if merged
        is_merged = False
        merge_range = None
        for merge in ws.merged_cells.ranges:
            if cell.coordinate in merge:
                is_merged = True
                merge_range = str(merge)
                # Only show details for top-left cell
                if merge.min_row == row_num and merge.min_col == col_num:
                    break
                else:
                    merge_range = f"part of {merge_range}"
                    break
        
        # Get formatting
        font = cell.font
        border = cell.border
        alignment = cell.alignment
        fill = cell.fill
        
        # Only print details if it's a top-left merged cell or unmerged cell with content
        if (is_merged and merge_range and not merge_range.startswith('part of')) or \
           (not is_merged and cell.value not in [None, '', 'None']):
            print(f"\n  {col_letter}{row_num}: '{cell.value}'")
            if is_merged:
                print(f"    MERGED: {merge_range}")
            print(f"    Font: {font.name}, Size: {font.size}, Bold: {font.bold}")
            print(f"    Alignment: H={alignment.horizontal}, V={alignment.vertical}, Wrap={alignment.wrap_text}")
            print(f"    Borders: Top={border.top.style if border.top else None}, "
                  f"Bottom={border.bottom.style if border.bottom else None}, "
                  f"Left={border.left.style if border.left else None}, "
                  f"Right={border.right.style if border.right else None}")
            if fill.patternType:
                print(f"    Fill: {fill.patternType}, Color: {fill.fgColor.rgb if fill.fgColor else 'None'}")

print("\n\n=== ALL MERGED RANGES IN ROWS 42-49 ===\n")
for merge in ws.merged_cells.ranges:
    if 42 <= merge.min_row <= 49:
        top_left = ws.cell(merge.min_row, merge.min_col)
        print(f"{merge}: '{top_left.value}'")
