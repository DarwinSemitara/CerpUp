from openpyxl import load_workbook
from openpyxl.cell import MergedCell

wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("=== DETAILED CELL-BY-CELL ANALYSIS (rows 26-31, cols A-K) ===\n")

for row_num in range(26, 32):  # rows 26-31
    print(f"ROW {row_num}:")
    for col_num in range(1, 12):  # cols 1-11 (A-K)
        col_letter = chr(64 + col_num)
        cell = ws.cell(row=row_num, column=col_num)
        
        is_merged = isinstance(cell, MergedCell)
        
        if is_merged:
            # Find which merge range this belongs to
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                    print(f"  {col_letter}{row_num}: MERGED (part of {chr(64+merged_range.min_col)}{merged_range.min_row}:{chr(64+merged_range.max_col)}{merged_range.max_row}) - value in top-left: '{top_left.value}'")
                    break
        else:
            # Regular cell
            print(f"  {col_letter}{row_num}: '{cell.value}'")
    print()
