from openpyxl import load_workbook

template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = template.active

print("\n=== BORDER DETAILS FOR ROW 30 (with underlines) ===\n")

cells = ['A30', 'B30', 'C30', 'D30', 'E30', 'F30', 'G30', 'H30', 'I30', 'J30', 'K30']

for cell_ref in cells:
    cell = ws[cell_ref]
    border = cell.border
    print(f"{cell_ref}:")
    print(f"  Top: {border.top.style if border.top else 'None'}")
    print(f"  Bottom: {border.bottom.style if border.bottom else 'None'}")
    print(f"  Left: {border.left.style if border.left else 'None'}")
    print(f"  Right: {border.right.style if border.right else 'None'}")
    print()
