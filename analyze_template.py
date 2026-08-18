from openpyxl import load_workbook

wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("=== CONCURRENT TEACHING SECTION (rows 26-31) ===\n")

# Check merged cells in this range
merges = []
for merged_range in ws.merged_cells.ranges:
    if 26 <= merged_range.min_row <= 31:
        merges.append(merged_range)

# Sort by row then column
merges.sort(key=lambda m: (m.min_row, m.min_col))

for merged_range in merges:
    print(f"Row {merged_range.min_row}, "
          f"Cols {merged_range.min_col}-{merged_range.max_col} "
          f"({chr(64+merged_range.min_col)}:{chr(64+merged_range.max_col)})")

    # Show cell value
    cell_val = ws.cell(merged_range.min_row, merged_range.min_col).value
    print(f"  Value: '{cell_val}'")
    print()
