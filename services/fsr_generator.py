"""
Faculty Service Record (FSR) Generator
Generates FSR Excel reports using FSRFORMAT.xlsx template with openpyxl

Architecture:
===========
- FSRFORMAT.xlsx is the SOURCE OF TRUTH for layout and formatting
- This code ONLY populates the template with data
- Template rows are used for copying formatting when expanding dynamic sections
- Never recreate formatting programmatically if it can be copied from template

Template Structure:
==================
Rows 1-10:   Header section (fixed)
Row 11:      Teaching load table header (fixed)
Rows 12-21:  Teaching load data rows (DYNAMIC - 10 template rows)
Row 22:      Teaching load TOTAL row (fixed)
Rows 23-25:  Teaching footnotes (fixed - 3 rows)
Rows 26-51:  Concurrent teaching section (fixed - 26 rows)
Rows 52-136: Research section (DYNAMIC - 85 template rows)
Rows 137+:   Extensions section (DYNAMIC)

Dynamic Section Handling:
========================
When a dynamic section needs more/fewer rows:
1. Calculate rows needed vs template rows
2. If more needed: insert rows after last template data row, copy formatting from template row
3. If fewer needed: delete excess rows (but preserve at least 1 for template reference)
4. Always preserve sections below the dynamic section
5. Track actual row positions after modifications

Critical Rules:
==============
1. NEVER delete the template row itself - keep it for formatting reference
2. NEVER hardcode formatting - copy from template
3. ALWAYS preserve merged cells - write to top-left cell only
4. ALWAYS update row tracking after insert/delete operations
5. NEVER overwrite sections below dynamic tables
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy
import os


# ══════════════════════════════════════════════════════════════
# Template Structure Configuration
# ══════════════════════════════════════════════════════════════

# These are INITIAL positions in the template file
# Actual positions will be tracked dynamically during generation
FSR_TEMPLATE = {
    # ── Fixed Header Section (rows 1-10) ──
    'semester_year': 'D2',          # Academic year and semester (merged D2:H2)
    'printed_name_last': 'C4',      # Last name (Family) (merged C4:D4)
    'printed_name_first': 'E4',     # First name (Given) (merged E4:F4)
    'printed_name_mi': 'G4',        # Middle initial
    'rank': 'I4',                   # Rank (merged I4:K4)
    'department': 'C7',             # Home Department (merged C7:H7)
    'college': 'I7',                # Home College (merged I7:K7)

    # ── Teaching Load Section (DYNAMIC) ──
    'teaching_header_row': 11,
    'teaching_template_row': 12,        # Template data row for copying formatting
    'teaching_data_start': 12,          # First data row
    # Last data row in template (10 rows: 12-21)
    'teaching_data_end': 21,
    'teaching_total_row': 22,           # TOTAL row (fixed relative position)

    # ── Teaching Footnotes Section (fixed - 3 rows) ──
    'footnotes_start': 23,
    'footnotes_end': 25,

    # ── Concurrent Teaching Section (fixed - 26 rows) ──
    'concurrent_start': 26,
    'concurrent_end': 51,

    # ── Research Section (DYNAMIC) ──
    'research_header_row': 52,          # Section header
    'research_template_row': 53,        # Template data row for copying formatting
    'research_data_start': 53,          # First data row
    'research_data_end': 136,           # Last data row in template (84 rows)

    # ── Extensions Section (DYNAMIC) ──
    'extensions_header_row': 137,       # Section header
    'extensions_template_row': 138,     # Template data row for copying formatting
    'extensions_data_start': 138,       # First data row
}


class FSRGenerator:
    """
    Generates Faculty Service Record Excel files using template-based approach.

    The template (FSRFORMAT.xlsx) controls ALL formatting and layout.
    This generator ONLY:
    1. Loads the template
    2. Populates fixed cells with data
    3. Expands/contracts dynamic sections based on data volume
    4. Copies formatting from template rows when adding new rows
    5. Preserves all content below dynamic sections
    """

    def __init__(self):
        self.template_path = 'static/reference/FSRFORMAT.xlsx'
        # Track actual row positions after modifications (updated during generation)
        self.positions = {}

    # ── Public entry points ───────────────────────────────────────────────

    def generate_fsr(self, faculty_data, research_data, extensions_data,
                     output_path, schedule_data=None, footnotes_data=None):
        """
        Generate an FSR Excel file using template-based approach.

        Args:
            faculty_data:    dict with name, rank, department, etc.
            research_data:   list of research project dicts
            extensions_data: list of extension activity dicts
            output_path:     path to save the generated .xlsx
            schedule_data:   list of schedule entry dicts (optional)

        Returns:
            output_path: Path to generated file
        """
        # Load template (never modify the original)
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active

        # Set worksheet title
        ws.title = f"({faculty_data.get('rank_number', '1')}) {faculty_data.get('last_name', 'Faculty')}"

        # Initialize position tracking with template defaults
        self.positions = FSR_TEMPLATE.copy()

        # Fill sections
        self._fill_header(ws, faculty_data)

        # Teaching load section (dynamic)
        self._fill_teaching_load(ws, schedule_data or [])

        # Footnotes section (dynamic - from database)
        self._fill_footnotes(ws, footnotes_data or [])

        # Research proposals section (dynamic)
        self._fill_research_proposals(ws, research_data or [])

        # DO NOT modify anything below research section yet
        # Extensions and other sections remain as-is in the template

        # Save to new file (never overwrite template)
        wb.save(output_path)
        return output_path

    def generate_fsr_for_member(self, member_id,
                                semester='2nd Semester',
                                academic_year='2025-2026'):
        """
        Generate FSR for a member by fetching all data from the database.
        Teaching load is matched by faculty last name in the schedules table.
        """
        from services.supabase_service import supabase

        # Fetch member from Supabase
        try:
            member_response = supabase.table('members').select(
                '*').eq('id', member_id).execute()  # Use 'id' not 'uid'
            if not member_response.data or len(member_response.data) == 0:
                raise ValueError(f"Member {member_id} not found")
            member_data = member_response.data[0]
        except Exception as e:
            raise ValueError(f"Member {member_id} not found: {e}")

        # Fetch research data
        research_data = []
        try:
            research_response = supabase.table('research').select(
                '*').eq('member_id', member_id).execute()
            research_data = research_response.data or []
        except Exception as e:
            print(f"Warning: could not fetch research data: {e}")

        # Fetch extensions data
        extensions_data = []
        try:
            extensions_response = supabase.table('extensions').select(
                '*').eq('member_id', member_id).execute()
            extensions_data = extensions_response.data or []
        except Exception as e:
            print(f"Warning: could not fetch extensions data: {e}")

        # Match schedules by last name and filter by semester/year
        last_name = (member_data.get('last') or '').strip().lower()
        schedule_data = []
        configured_subjects_data = []

        try:
            # Extract semester number (e.g., "1st Semester" -> "1")
            sem_num = semester.split()[0][0]  # Get first character

            # Fetch schedules
            rows = supabase.table('schedules').select('*').execute().data or []
            for s in rows:
                prof = (s.get('prof') or '').lower()
                # Match by last name, school year, and semester
                year_match = s.get('school_year') == academic_year
                sem_match = s.get('semester') == sem_num

                if last_name and last_name in prof and year_match and sem_match:
                    schedule_data.append({
                        'subjCode': s.get('subj_code') or s.get('subjCode', ''),
                        'subjName': s.get('subj_name') or s.get('subjName', ''),
                        'room':     s.get('room', ''),
                        'day':      s.get('day', ''),
                        'start':    s.get('start', ''),
                        'end':      s.get('end', ''),
                        'section':  s.get('section', ''),
                        'units':    s.get('units', ''),
                    })

            # Fetch configured subjects (TBA entries)
            configured_response = supabase.table('configured_subjects').select(
                '*').eq('school_year', academic_year).eq('semester', sem_num).execute()
            for cs in (configured_response.data or []):
                cs_prof = (cs.get('prof') or '').lower()
                if last_name and last_name in cs_prof:
                    # Check if this subject is already scheduled
                    is_scheduled = any(
                        sch['subjCode'] == cs.get('subj_code')
                        for sch in schedule_data
                    )
                    if not is_scheduled:
                        # Add as TBA entry
                        configured_subjects_data.append({
                            'subjCode': cs.get('subj_code', ''),
                            'subjName': cs.get('subj_name', ''),
                            'room':     'TBA',
                            'day':      'TBA',
                            'start':    '',
                            'end':      '',
                            'section':  cs.get('section', ''),
                            'units':    cs.get('units', ''),
                        })
        except Exception as e:
            print(f"Warning: could not fetch schedule data: {e}")

        # Combine scheduled and configured (TBA) subjects
        all_schedule_data = schedule_data + configured_subjects_data

        # Fetch footnotes from Supabase
        footnotes_data = []
        try:
            # Extract semester number (e.g., "1st Semester" -> "1")
            sem_num = semester.split()[0][0]  # Get first character

            # Get footnotes directly for this member/semester/year
            footnotes_result = supabase.table('fsr_footnotes').select(
                '*'
            ).eq('member_id', member_id).eq(
                'semester', sem_num
            ).eq('academic_year', academic_year).order('footnote_number').execute()

            # Format footnotes for FSR
            for fn in (footnotes_result.data or []):
                footnotes_data.append({
                    'footnote_number': fn.get('footnote_number', 1),
                    'footnote_type': fn.get('footnote_type', 'team'),
                    'faculty_name': fn.get('faculty_name', ''),
                    'subject': fn.get('subject', '')
                })

            print(
                f"📝 Fetched {len(footnotes_data)} footnotes for member {member_id}, semester {sem_num}, year {academic_year}")
            if footnotes_data:
                print(f"   Footnotes: {footnotes_data}")
        except Exception as e:
            print(f"Warning: Could not fetch footnotes: {e}")
            footnotes_data = []

        faculty_data = {
            'last_name':       member_data.get('last', ''),
            'first_name':      member_data.get('first', ''),
            'middle_initial':  member_data.get('middle', ''),
            'rank':            member_data.get('rank', 'N/A'),
            'rank_number':     member_data.get('rank_number', '1'),
            'department':      member_data.get('department', 'DCERP'),
            'college':         member_data.get('college', 'CHE'),
            'employment_type': member_data.get('employment_type', 'full_time'),
            'semester':        semester,
            'academic_year':   academic_year,
        }

        output_dir = 'generated_fsr'
        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_output_path = os.path.join(output_dir,
                                        f"FSR_{member_data.get('last', 'Faculty')}_{ts}.xlsx")

        # Generate FSR to temporary local file
        self.generate_fsr(faculty_data, [], [],
                          temp_output_path, schedule_data=all_schedule_data,
                          footnotes_data=footnotes_data)

        # Upload to Supabase Storage and save metadata
        try:
            file_metadata = self._upload_fsr_to_storage(
                temp_output_path,
                member_id,
                member_data,
                semester,
                academic_year
            )

            # Delete local file after successful upload
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
                print(f"✓ Deleted local file: {temp_output_path}")

            return file_metadata
        except Exception as e:
            print(f"Error uploading FSR to storage: {e}")
            # Return local path as fallback
            return {'local_path': temp_output_path, 'error': str(e)}

    def _upload_fsr_to_storage(self, file_path, member_id, member_data, semester, academic_year):
        """
        Upload FSR file to Supabase Storage and save metadata to database.

        Returns:
            dict: File metadata with download URL
        """
        from services.supabase_service import supabase

        # Read file
        with open(file_path, 'rb') as f:
            file_content = f.read()

        # Get file info
        file_size = os.path.getsize(file_path)
        file_name = os.path.basename(file_path)

        # Create storage path: academic_year/semester/filename
        # Clean up names for path
        year_folder = academic_year.replace('/', '-')
        semester_folder = semester.replace(' ', '-')
        storage_path = f"{year_folder}/{semester_folder}/{file_name}"

        # Upload to Supabase Storage
        bucket_name = 'fsr-files'

        try:
            # Upload file (will overwrite if exists)
            supabase.storage.from_(bucket_name).upload(
                storage_path,
                file_content,
                file_options={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
            )
            print(f"✓ Uploaded to storage: {storage_path}")
        except Exception as upload_error:
            # If file exists, update it
            if 'already exists' in str(upload_error).lower():
                supabase.storage.from_(bucket_name).update(
                    storage_path,
                    file_content,
                    file_options={
                        "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
                )
                print(f"✓ Updated existing file: {storage_path}")
            else:
                raise

        # Get public URL
        public_url = supabase.storage.from_(
            bucket_name).get_public_url(storage_path)

        # Save metadata to database
        member_name = f"{member_data.get('first', '')} {member_data.get('last', '')}".strip(
        )
        member_email = member_data.get('email', '')

        # Check if record exists (upsert logic)
        existing = supabase.table('fsr_files').select('id').eq(
            'member_id', member_id
        ).eq('semester', semester).eq('academic_year', academic_year).is_(
            'deleted_at', None
        ).execute()

        file_metadata = {
            'member_id': member_id,
            'member_name': member_name,
            'member_email': member_email,
            'semester': semester,
            'academic_year': academic_year,
            'file_path': storage_path,
            'file_name': file_name,
            'file_size': file_size,
            'storage_bucket': bucket_name,
        }

        if existing.data and len(existing.data) > 0:
            # Update existing record
            record_id = existing.data[0]['id']
            supabase.table('fsr_files').update(
                file_metadata).eq('id', record_id).execute()
            print(f"✓ Updated FSR metadata: {record_id}")
        else:
            # Insert new record
            result = supabase.table('fsr_files').insert(
                file_metadata).execute()
            record_id = result.data[0]['id'] if result.data else None
            print(f"✓ Saved FSR metadata: {record_id}")

        return {
            'id': record_id,
            'file_name': file_name,
            'download_url': public_url,
            'file_size': file_size,
            'storage_path': storage_path
        }

    # ── Private helpers ───────────────────────────────────────────────────

    def _write_cell(self, ws, cell_ref, value):
        """
        Write to a cell safely.
        For merged cells, writes to the top-left cell without unmerging.

        Args:
            ws: Worksheet object
            cell_ref: Cell reference like 'A1', 'E2'
            value: Value to write
        """
        from openpyxl.cell import MergedCell

        cell = ws[cell_ref]

        # If it's a merged cell, find the top-left cell of the merge range
        if isinstance(cell, MergedCell):
            # Get the merged range that contains this cell
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    # Write to the top-left cell of the merged range
                    top_left = merged_range.start_cell
                    ws[top_left.coordinate].value = value
                    return

        # Not a merged cell, write directly
        cell.value = value

    def _copy_row_style(self, ws, source_row, target_row, start_col=1, end_col=None):
        """
        Copy ALL styling from source row to target row.
        This is critical for maintaining template formatting.

        Args:
            ws: Worksheet object
            source_row: Source row number
            target_row: Target row number
            start_col: Starting column (default 1 = A)
            end_col: Ending column (default None = max column)
        """
        if end_col is None:
            end_col = ws.max_column

        for col in range(start_col, end_col + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)

            # Copy all style attributes if source has style
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            # Copy row height from source to target
            if ws.row_dimensions[source_row].height:
                ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    def _fill_header(self, ws, faculty_data):
        """
        Fill the fixed header section (rows 1-10).
        Uses FSR_TEMPLATE cell mapping for merged cells.
        """
        semester = faculty_data.get('semester', '2nd Semester')
        acad_year = faculty_data.get('academic_year', '2025-2026')

        # D2:H2 is merged, write to top-left cell D2
        self._write_cell(ws, FSR_TEMPLATE['semester_year'],
                         f"{semester} {acad_year}")

        # Row 4: Printed name (Last, First, MI) - all merged cells
        last = faculty_data.get('last_name', '').upper()
        first = faculty_data.get('first_name', '').upper()
        mi = faculty_data.get('middle_initial', '').upper()

        self._write_cell(ws, FSR_TEMPLATE['printed_name_last'], last)
        self._write_cell(ws, FSR_TEMPLATE['printed_name_first'], first)
        self._write_cell(ws, FSR_TEMPLATE['printed_name_mi'], mi)
        self._write_cell(ws, FSR_TEMPLATE['rank'], 'N/A')

        # Row 7: Department and College - both merged cells
        self._write_cell(ws, FSR_TEMPLATE['department'],
                         faculty_data.get('department', 'DCERP'))
        self._write_cell(ws, FSR_TEMPLATE['college'],
                         faculty_data.get('college', 'CHE'))

    def _fill_teaching_load(self, ws, schedule_data):
        """
        Fill Section I Teaching Load (DYNAMIC section).

        Template structure:
        - Row 11: Header (fixed)
        - Rows 12-21: Data rows (10 template rows)
        - Row 22: TOTAL row (fixed)

        Process:
        1. Consolidate schedule data by subject/room/time
        2. Calculate rows needed vs template rows (10)
        3. Insert or delete rows as needed
        4. Copy formatting from template row 12
        5. Populate with data
        6. Preserve TOTAL row and all sections below

        Returns:
            int: Number of rows added (positive) or removed (negative)
        """
        TEMPLATE_ROW = self.positions['teaching_template_row']  # Row 12
        DATA_START = self.positions['teaching_data_start']      # Row 12
        DATA_END = self.positions['teaching_data_end']          # Row 21
        TOTAL_ROW = self.positions['teaching_total_row']        # Row 22
        TEMPLATE_DATA_ROWS = DATA_END - DATA_START + 1          # 10 rows
        ORIGINAL_DATA_END = 21  # Always use original template end row for clearing

        # Day abbreviations mapping
        DAY_ABBR = {
            'Monday': 'M', 'Tuesday': 'T', 'Wednesday': 'W',
            'Thursday': 'TH', 'Friday': 'F', 'Saturday': 'S',
        }

        def fmt_time(start, end):
            """Format time range like '7:00-8:30 AM' or '4:00-5:30 PM'"""
            if not start or not end:
                return 'TBA'
            try:
                hs, ms = int(start.split(':')[0]), int(start.split(':')[1])
                he, me = int(end.split(':')[0]), int(end.split(':')[1])
                ps = 'PM' if hs >= 12 else 'AM'
                pe = 'PM' if he >= 12 else 'AM'
                hs12 = hs - 12 if hs > 12 else (12 if hs == 0 else hs)
                he12 = he - 12 if he > 12 else (12 if he == 0 else he)
                if ps == pe:
                    return f"{hs12}:{ms:02d}-{he12}:{me:02d} {pe}"
                return f"{hs12}:{ms:02d} {ps}-{he12}:{me:02d} {pe}"
            except Exception:
                return f"{start}-{end}"

        # Consolidate schedules: same subject+room+time → merge days
        consolidated = {}
        for s in sorted(schedule_data, key=lambda x: (x.get('subjCode', ''), x.get('day', ''))):
            key = (s.get('subjCode', ''), s.get('room', ''),
                   s.get('start', ''), s.get('end', ''))
            if key not in consolidated:
                consolidated[key] = {'entry': s, 'days': []}
            day = s.get('day', '')
            abbr = DAY_ABBR.get(day, (day or '')[:2].upper() if day else '')
            if abbr and abbr not in consolidated[key]['days']:
                consolidated[key]['days'].append(abbr)

        rows_to_write = list(consolidated.values())
        num_subjects = len(rows_to_write)

        # Handle zero subjects case - keep 1 empty template row
        if num_subjects == 0:
            num_subjects = 1
            rows_to_write = [{'entry': {}, 'days': []}]

        # Calculate row difference
        rows_difference = num_subjects - TEMPLATE_DATA_ROWS
        rows_added = 0

        # FIRST: Clear ALL 10 template data rows BEFORE any insertions/deletions
        print(
            f"🧹 Clearing template rows {DATA_START} to {DATA_END} (before row modifications)")
        for clear_row in range(DATA_START, DATA_END + 1):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                self._write_cell(ws, f"{col}{clear_row}", None)
        print(f"✓ Template rows cleared")

        # THEN: Adjust row count if needed
        if rows_difference > 0:
            # Need MORE rows - insert after the last template data row (row 21)
            insert_position = DATA_END + 1
            for i in range(rows_difference):
                ws.insert_rows(insert_position)
                # Copy formatting from template row (row 12)
                self._copy_row_style(
                    ws, TEMPLATE_ROW, insert_position, start_col=1, end_col=11)
                insert_position += 1
            rows_added = rows_difference

            # UPDATE: Adjust TOTAL row position after insertion
            self.positions['teaching_total_row'] = TOTAL_ROW + rows_difference

        elif rows_difference < 0:
            # Need FEWER rows - delete them and manually restore concurrent teaching merges
            delete_start = DATA_END
            delete_count = abs(rows_difference)

            print(
                f"🗑️ Deleting {delete_count} rows from {delete_start - delete_count + 1} to {delete_start}")

            # BEFORE DELETION: Save ALL merge structures from row 23 onwards
            # This includes footnotes (23-25) + concurrent teaching (26-51) + rest
            below_teaching_start = 23
            merges_to_restore = []

            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row >= below_teaching_start:
                    merges_to_restore.append({
                        'min_row': merged_range.min_row,
                        'max_row': merged_range.max_row,
                        'min_col': merged_range.min_col,
                        'max_col': merged_range.max_col
                    })

            print(
                f"💾 Saved {len(merges_to_restore)} merge structures from row {below_teaching_start} onwards")

            # PERFORM DELETION
            ws.delete_rows(delete_start - delete_count + 1, delete_count)
            rows_added = rows_difference  # Negative value (e.g., -7)

            # AFTER DELETION: Concurrent teaching section (27-41) gets corrupted
            # This is 15 rows total with complex merge structure
            # Row 27 shifts to 20 (with -7 shift)
            r_base = 27 + rows_difference

            print(
                f"   🔧 Recreating concurrent teaching at rows {r_base}-{r_base+14}")

            # Save font styles from template BEFORE they get corrupted
            from copy import copy
            template_wb = load_workbook(self.template_path)
            template_ws = template_wb.active

            font_a27 = copy(template_ws['A27'].font)
            font_a30 = copy(template_ws['A30'].font)
            font_e30 = copy(template_ws['E30'].font)
            font_h30 = copy(template_ws['H30'].font)
            font_a31 = copy(template_ws['A31'].font)
            font_e31 = copy(template_ws['E31'].font)
            font_h31 = copy(template_ws['H31'].font)
            font_a34 = copy(template_ws['A34'].font)
            font_a35 = copy(template_ws['A35'].font)

            # Also save border styles
            border_a30 = copy(template_ws['A30'].border)
            border_e30 = copy(template_ws['E30'].border)
            border_h30 = copy(template_ws['H30'].border)
            border_a31 = copy(template_ws['A31'].border)
            border_e31 = copy(template_ws['E31'].border)
            border_h31 = copy(template_ws['H31'].border)

            # Unmerge any broken merges in this range
            for row in range(r_base, r_base+15):
                try:
                    ws.unmerge_cells(f"A{row}:K{row}")
                except:
                    pass

            # Row 27→r_base: A27:K27 - "Concurrent teaching load..."
            ws.merge_cells(start_row=r_base, start_column=1,
                           end_row=r_base, end_column=11)
            self._write_cell(
                ws, f'A{r_base}', 'Concurrent teaching load outside the college.   Write NONE whenever applicable.  Please do not leave any blank')

            # Row 28→r_base+1: A28:K28 - " "
            ws.merge_cells(start_row=r_base+1, start_column=1,
                           end_row=r_base+1, end_column=11)
            self._write_cell(ws, f'A{r_base+1}', ' ')

            # Row 29→r_base+2: No merges, all "None"

            # Row 30→r_base+3: A30:C30, E30:F30, H30:K30 - "(NONE)"
            ws.merge_cells(start_row=r_base+3, start_column=1,
                           end_row=r_base+3, end_column=3)
            self._write_cell(ws, f'A{r_base+3}', '(NONE)')
            ws.merge_cells(start_row=r_base+3, start_column=5,
                           end_row=r_base+3, end_column=6)
            self._write_cell(ws, f'E{r_base+3}', '(NONE)')
            ws[f'E{r_base+3}'].font = font_e30
            ws[f'E{r_base+3}'].border = border_e30
            ws.merge_cells(start_row=r_base+3, start_column=8,
                           end_row=r_base+3, end_column=11)
            self._write_cell(ws, f'H{r_base+3}', '(NONE)')
            ws[f'H{r_base+3}'].font = font_h30
            ws[f'H{r_base+3}'].border = border_h30

            # Row 31→r_base+4: A31, E31:F31, H31:K31
            self._write_cell(ws, f'A{r_base+4}', 'COLLEGE OUTSIDE U.P. SYSTEM')
            ws.merge_cells(start_row=r_base+4, start_column=5,
                           end_row=r_base+4, end_column=6)
            self._write_cell(ws, f'E{r_base+4}', 'No. of subjects')
            ws[f'E{r_base+4}'].font = font_e31
            ws[f'E{r_base+4}'].border = border_e31
            ws.merge_cells(start_row=r_base+4, start_column=8,
                           end_row=r_base+4, end_column=11)
            self._write_cell(ws, f'H{r_base+4}',
                             'No. of units (w/o multipliers)')
            ws[f'H{r_base+4}'].font = font_h31
            ws[f'H{r_base+4}'].border = border_h31

            # Row 32→r_base+5: No merges, all "None"

            # Row 33→r_base+6: A33:C33, E33:F33, H33:K33 - "(NONE)"
            ws.merge_cells(start_row=r_base+6, start_column=1,
                           end_row=r_base+6, end_column=3)
            self._write_cell(ws, f'A{r_base+6}', '(NONE)')
            ws.merge_cells(start_row=r_base+6, start_column=5,
                           end_row=r_base+6, end_column=6)
            self._write_cell(ws, f'E{r_base+6}', '(NONE)')
            ws[f'E{r_base+6}'].font = font_e30
            ws[f'E{r_base+6}'].border = border_e30
            ws.merge_cells(start_row=r_base+6, start_column=8,
                           end_row=r_base+6, end_column=11)
            self._write_cell(ws, f'H{r_base+6}', '(NONE)')
            ws[f'H{r_base+6}'].font = font_h30
            ws[f'H{r_base+6}'].border = border_h30

            # Row 34→r_base+7: A34:C34, E34:F34, H34:K34
            ws.merge_cells(start_row=r_base+7, start_column=1,
                           end_row=r_base+7, end_column=3)
            self._write_cell(ws, f'A{r_base+7}', 'U.P. COLLEGE/DEPT.')
            ws.merge_cells(start_row=r_base+7, start_column=5,
                           end_row=r_base+7, end_column=6)
            self._write_cell(ws, f'E{r_base+7}', 'No. of subjects')
            ws[f'E{r_base+7}'].font = font_e31
            ws[f'E{r_base+7}'].border = border_e31
            ws.merge_cells(start_row=r_base+7, start_column=8,
                           end_row=r_base+7, end_column=11)
            self._write_cell(ws, f'H{r_base+7}',
                             'No. of units (w/o multipliers)')
            ws[f'H{r_base+7}'].font = font_h31
            ws[f'H{r_base+7}'].border = border_h31

            # Row 35→r_base+8: A35:K35 - NOTE text
            ws.merge_cells(start_row=r_base+8, start_column=1,
                           end_row=r_base+8, end_column=11)
            self._write_cell(ws, f'A{r_base+8}', 'NOTE:   A faculty member teaching in another college and/or another Constituent University (CU) should file a separate  Form 67 (FSR) in that college and/or CU. In the event that a 2nd or 3rd FSR needs to be prepared, only the teaching load and consultation hours should be  completed. Permission fromf the Chancellor should be sought before teaching outside the University.')

            # Row 36→r_base+9: H36 only
            self._write_cell(ws, f'H{r_base+9}', 'Certified Correct:')

            # Row 37→r_base+10: No merges
            # Row 38→r_base+11: No merges

            # Row 39→r_base+12: G39:J39
            ws.merge_cells(start_row=r_base+12, start_column=7,
                           end_row=r_base+12, end_column=10)
            self._write_cell(ws, f'G{r_base+12}',
                             'MARGARITA CARMEN S. PATERNO')

            # Row 40→r_base+13: G40:J40
            ws.merge_cells(start_row=r_base+13, start_column=7,
                           end_row=r_base+13, end_column=10)
            self._write_cell(ws, f'G{r_base+13}', 'University Registrar')

            # Row 41→r_base+14: No merges

            # Restore other merges (skip rows 27-41)
            restored_count = 0
            for merge in merges_to_restore:
                new_min_row = merge['min_row'] + rows_difference
                if new_min_row < 1 or (r_base <= new_min_row <= r_base+14):
                    continue
                try:
                    ws.merge_cells(
                        start_row=new_min_row,
                        start_column=merge['min_col'],
                        end_row=new_min_row +
                        (merge['max_row'] - merge['min_row']),
                        end_column=merge['max_col']
                    )
                    restored_count += 1
                except:
                    pass

            print(f"✅ Restored {restored_count} other merges")

            # Adjust TOTAL row position
            self.positions['teaching_total_row'] = TOTAL_ROW + rows_difference

        # Now populate with actual data
        print(
            f"📝 Populating {num_subjects} subjects starting at row {DATA_START}")
        for i, item in enumerate(rows_to_write):
            row = DATA_START + i
            entry = item['entry']
            days = '/'.join(item['days']) if item['days'] else 'TBA'
            time = fmt_time(entry.get('start', ''), entry.get('end', ''))

            # Only populate if we have actual data
            if entry:
                subjCode = entry.get('subjCode', '')
                section = entry.get('section', '')
                room = entry.get('room', 'TBA')

                print(
                    f"  Row {row}: {subjCode} - {section} - {room} - {days} - {time}")

                # Column A: Subject Code
                self._write_cell(ws, f"A{row}", subjCode)
                # Column B: Section Code
                self._write_cell(ws, f"B{row}", section)
                # Column C: Room
                self._write_cell(ws, f"C{row}", room)
                # Column D: Days
                self._write_cell(ws, f"D{row}", days)
                # Column E: Time
                self._write_cell(ws, f"E{row}", time)

                # Column F: Hrs/Week - leave template default (member can edit)

                # Columns G, H, I, J, K: Set to 0 as requested
                self._write_cell(ws, f"G{row}", 0)  # TL
                self._write_cell(ws, f"H{row}", 0)  # OL
                self._write_cell(ws, f"I{row}", 0)  # RL
                self._write_cell(ws, f"J{row}", 0)  # No. of Students
                self._write_cell(ws, f"K{row}", 0)  # Teaching Load Credits

        # Update TOTAL row position in tracking
        self.positions['teaching_total_row'] = DATA_START + num_subjects

        return rows_added

    def _fill_footnotes(self, ws, footnotes_data):
        """
        Fill footnotes section dynamically after the TOTAL row.
        Footnotes come from the database (fsr_footnotes table).

        Args:
            ws: Worksheet object
            footnotes_data: List of footnote dicts from database with keys:
                           footnote_number, footnote_type, faculty_name, subject
        """
        print(f"🔖 _fill_footnotes called with {len(footnotes_data)} footnotes")

        # Footnotes start right after TOTAL row - ALWAYS 3 rows
        total_row = self.positions['teaching_total_row']
        footnote_start = total_row + 1
        footnote_end = footnote_start + 2  # Always exactly 3 rows

        print(
            f"   TOTAL row: {total_row}, clearing footnote rows: {footnote_start} to {footnote_end}")

        # FIRST: Clear the 3 template footnote rows (remove sample data)
        # Footnotes are merged cells A:K - just clear the top-left cell (A) without unmerging
        print(
            f"🧹 Clearing template footnote rows {footnote_start} to {footnote_end}")
        for clear_row in range(footnote_start, footnote_end + 1):
            # Use _write_cell to handle merged cells properly - preserves formatting
            self._write_cell(ws, f"A{clear_row}", "")  # Empty string, not None
        print(f"✓ Template footnote rows cleared")

        # THEN: Write actual footnotes (if any)
        if not footnotes_data:
            print("   No footnotes to write - leaving blank")
            return

        # Unicode footnote symbols
        footnote_symbols = {
            1: '¹', 2: '²', 3: '³', 4: '⁴', 5: '⁵',
            6: '⁶', 7: '⁷', 8: '⁸', 9: '⁹', 10: '¹⁰'
        }

        # Write each footnote to column A
        for footnote in footnotes_data:
            number = footnote.get('footnote_number', 1)
            ftype = footnote.get('footnote_type', 'team')
            faculty = footnote.get('faculty_name', '')
            subject = footnote.get('subject', '')

            symbol = footnote_symbols.get(number, str(number))

            # Format type text
            if ftype == 'relay':
                type_text = 'Relay teaching'
            else:
                type_text = 'Team teaching'

            # Format: "¹Team teaching with John Doe (CE 101)"
            if subject:
                text = f"{symbol}{type_text} with {faculty} ({subject})"
            else:
                text = f"{symbol}{type_text} with {faculty}"

            row = footnote_start + (number - 1)
            print(f"   Writing footnote {number} to A{row}: {text}")
            # Use _write_cell to handle merged cells properly
            self._write_cell(ws, f"A{row}", text)

    def _fill_research_proposals(self, ws, research_data):
        """
        Fill research proposals section dynamically (rows 42-49).

        Template structure:
        - Row 42: Section header "II. RESEARCH/TEXTBOOK WRITING/CREATIVE WORK:"
        - Row 43: Subsection "II.A RESEARCH"
        - Row 44: Sub-subsection "II.A1 RESEARCH PROPOSAL"
        - Row 45: Table headers
        - Rows 46-47: 2 template data rows
        - Row 48: Total row with formula
        - Row 49: Empty

        Args:
            ws: Worksheet object
            research_data: List of research dicts from database
        """
        print(
            f"📚 _fill_research_proposals called with {len(research_data)} research items")

        # Filter for proposals only
        proposals = [r for r in research_data if r.get(
            'research_type') == 'proposal']
        print(f"   Found {len(proposals)} research proposals")

        # Template positions (before any modifications)
        HEADER_ROW = 42  # "II. RESEARCH/TEXTBOOK WRITING/CREATIVE WORK:"
        SUBSECTION_ROW = 43  # "II.A RESEARCH"
        SUBSUBSECTION_ROW = 44  # "II.A1 RESEARCH PROPOSAL"
        TABLE_HEADER_ROW = 45
        DATA_START_ROW = 46  # First data row
        TEMPLATE_DATA_ROWS = 2  # Rows 46-47
        TOTAL_ROW = 48  # Total row

        num_proposals = len(proposals)

        # Calculate row adjustment needed
        rows_difference = num_proposals - TEMPLATE_DATA_ROWS

        if rows_difference == 0:
            print("   No row adjustment needed - exact match with template")
        elif rows_difference > 0:
            print(f"   Need to INSERT {rows_difference} rows")
        else:
            print(f"   Need to DELETE {abs(rows_difference)} rows")

        # STEP 1: Clear template data rows (46-47)
        print(
            f"🧹 Clearing template data rows {DATA_START_ROW} to {DATA_START_ROW+TEMPLATE_DATA_ROWS-1}")
        for clear_row in range(DATA_START_ROW, DATA_START_ROW + TEMPLATE_DATA_ROWS):
            # Clear each cell while preserving formatting
            self._write_cell(ws, f"A{clear_row}", "")
            self._write_cell(ws, f"E{clear_row}", "")
            self._write_cell(ws, f"F{clear_row}", "")
            self._write_cell(ws, f"I{clear_row}", "")
            self._write_cell(ws, f"K{clear_row}", "")
        print("✓ Template data rows cleared")

        # STEP 2: Adjust rows if needed
        if rows_difference > 0:
            # Insert additional rows
            print(f"➕ Inserting {rows_difference} rows at row {TOTAL_ROW}")
            ws.insert_rows(TOTAL_ROW, rows_difference)

            # Copy formatting from row 47 to new rows
            from copy import copy
            for offset in range(rows_difference):
                new_row = TOTAL_ROW + offset
                # Copy cell formatting from row 47
                for col_letter in ['A', 'E', 'F', 'I', 'K']:
                    template_cell = ws[f"{col_letter}47"]
                    new_cell = ws[f"{col_letter}{new_row}"]
                    new_cell.font = copy(template_cell.font)
                    new_cell.border = copy(template_cell.border)
                    new_cell.alignment = copy(template_cell.alignment)
                    new_cell.fill = copy(template_cell.fill)

            # Recreate merged cells in new rows
            for offset in range(rows_difference):
                new_row = TOTAL_ROW + offset
                ws.merge_cells(f"A{new_row}:D{new_row}")
                ws.merge_cells(f"F{new_row}:H{new_row}")
                ws.merge_cells(f"I{new_row}:J{new_row}")

            print(f"✓ Inserted {rows_difference} rows with formatting")

        elif rows_difference < 0:
            # Delete excess rows
            delete_count = abs(rows_difference)
            delete_start = TOTAL_ROW - delete_count
            print(f"➖ Deleting {delete_count} rows from {delete_start}")
            ws.delete_rows(delete_start, delete_count)
            print(f"✓ Deleted {delete_count} rows")

        # STEP 3: Populate with actual proposal data
        print(
            f"📝 Populating {num_proposals} proposals starting at row {DATA_START_ROW}")
        for i, proposal in enumerate(proposals):
            row = DATA_START_ROW + i

            # Write data to merged cells
            self._write_cell(ws, f"A{row}", proposal.get('title', ''))
            self._write_cell(ws, f"E{row}", proposal.get('role', ''))
            self._write_cell(ws, f"F{row}", proposal.get('co_workers', ''))
            self._write_cell(ws, f"I{row}", proposal.get('funding_agency', ''))
            self._write_cell(ws, f"K{row}", proposal.get('credit_units', ''))

            print(f"  Row {row}: {proposal.get('title', '')[:50]}")

        # STEP 4: Update TOTAL formula
        new_total_row = DATA_START_ROW + num_proposals
        if num_proposals > 0:
            formula = f"=SUM(K{DATA_START_ROW}:K{DATA_START_ROW+num_proposals-1})"
        else:
            formula = "0"
        ws[f"K{new_total_row}"].value = formula
        print(f"✓ Updated TOTAL formula at K{new_total_row}: {formula}")


# ── Convenience function ──────────────────────────────────────────────────────

def generate_member_fsr(member_id, semester='2nd Semester', academic_year='2025-2026'):
    """Generate FSR for a member — thin wrapper around FSRGenerator."""
    return FSRGenerator().generate_fsr_for_member(member_id, semester, academic_year)
