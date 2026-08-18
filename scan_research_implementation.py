from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load template
wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("=" * 100)
print("RESEARCH IMPLEMENTATION SECTION ANALYSIS")
print("=" * 100)

# Based on the template, Research Implementation should start around row 48-56
# Let's scan from row 42 to 60 to find the section
start_scan = 42
end_scan = 60

print(f"\n📊 Scanning rows {start_scan} to {end_scan} for Research Implementation section...\n")

# Find the section header
for row in range(start_scan, end_scan + 1):
    cell_value = ws[f'A{row}'].value
    if cell_value and 'RESEARCH IMPLEMENTATION' in str(cell_value).upper():
        print(f"✅ Found section header at row {row}: {cell_value}")
        section_start = row
        break
else:
    section_start = 48  # Default guess

# Analyze structure from section start
print(f"\n📋 SECTION STRUCTURE (starting from row {section_start}):\n")

for row in range(section_start, section_start + 12):
    print(f"Row {row}:")
    
    # Show merged cells in this row
    merged_in_row = []
    for merge in ws.merged_cells.ranges:
        if merge.min_row == row:
            merged_in_row.append(str(merge))
    
    if merged_in_row:
        print(f"  Merges: {', '.join(merged_in_row)}")
    
    # Show content in key columns
    content = []
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        cell = ws[f'{col}{row}']
        if cell.value:
            content.append(f"{col}='{str(cell.value)[:30]}'")
    
    if content:
        print(f"  Content: {', '.join(content)}")
    
    # Check if this is a data row (has sample data)
    a_cell = ws[f'A{row}'].value
    if a_cell and ('Proposal' in str(a_cell) or 'SAMPLE' in str(a_cell).upper() or 'Lorem' in str(a_cell)):
        print(f"  ⭐ DATA ROW - contains sample data")
    
    print()

# Identify key rows
print("\n" + "=" * 100)
print("KEY ROW IDENTIFICATION")
print("=" * 100)

header_row = None
data_start_row = None
data_rows = []
total_row = None

for row in range(section_start, section_start + 12):
    cell_a = ws[f'A{row}'].value
    cell_k = ws[f'K{row}'].value
    
    # Header row has "TITLE (SPECIFY COMPLETE TITLE)"
    if cell_a and 'TITLE' in str(cell_a).upper() and 'COMPLETE' in str(cell_a).upper():
        header_row = row
        print(f"📌 HEADER ROW: {row}")
    
    # Data rows contain sample data
    if cell_a and any(keyword in str(cell_a) for keyword in ['Proposal', 'SAMPLE', 'Lorem', 'ipsum']):
        data_rows.append(row)
        if data_start_row is None:
            data_start_row = row
        print(f"📝 DATA ROW: {row}")
    
    # Total row has "Total Research"
    if cell_a and 'Total Research' in str(cell_a):
        total_row = row
        print(f"🔢 TOTAL ROW: {row}")

print(f"\nTemplate has {len(data_rows)} sample data rows")

# Analyze data row structure in detail
if data_rows:
    print("\n" + "=" * 100)
    print("DATA ROW STRUCTURE (Detailed)")
    print("=" * 100)
    
    sample_row = data_rows[0]
    print(f"\nAnalyzing row {sample_row} as template:\n")
    
    # Check each column
    for col_idx in range(1, 12):  # A to K
        col = get_column_letter(col_idx)
        cell = ws[f'{col}{sample_row}']
        
        # Check if part of merge
        merge_info = "standalone"
        for merge in ws.merged_cells.ranges:
            if cell.coordinate in merge:
                merge_info = f"part of {merge}"
                break
        
        # Border info
        border_info = []
        if cell.border:
            if cell.border.left and cell.border.left.style:
                border_info.append(f"L:{cell.border.left.style}")
            if cell.border.right and cell.border.right.style:
                border_info.append(f"R:{cell.border.right.style}")
            if cell.border.top and cell.border.top.style:
                border_info.append(f"T:{cell.border.top.style}")
            if cell.border.bottom and cell.border.bottom.style:
                border_info.append(f"B:{cell.border.bottom.style}")
        
        border_str = ", ".join(border_info) if border_info else "no borders"
        
        value = str(cell.value)[:40] if cell.value else "(empty)"
        
        print(f"  {col}{sample_row}: {merge_info:25} | Borders: {border_str:40} | Value: {value}")

print("\n" + "=" * 100)
print("SUMMARY")
print("=" * 100)
print(f"Section starts at row: {section_start}")
print(f"Header row: {header_row}")
print(f"Data starts at row: {data_start_row}")
print(f"Template data rows: {data_rows}")
print(f"Total row: {total_row}")
print(f"Number of template data rows: {len(data_rows)}")
