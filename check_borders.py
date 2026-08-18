from openpyxl import load_workbook

wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("Template row 46 borders:")
for col in ['I', 'J', 'K']:
    cell = ws[f'{col}46']
    print(f"  {col}46: left={cell.border.left.style if cell.border and cell.border.left else None}, "
          f"right={cell.border.right.style if cell.border and cell.border.right else None}")

print("\nTemplate row 47 borders:")
for col in ['I', 'J', 'K']:
    cell = ws[f'{col}47']
    print(f"  {col}47: left={cell.border.left.style if cell.border and cell.border.left else None}, "
          f"right={cell.border.right.style if cell.border and cell.border.right else None}")
