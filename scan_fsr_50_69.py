from openpyxl import load_workbook

wb = load_workbook('static/reference/FSRFORMAT.xlsx')
ws = wb.active

print("ROW SCAN 50-69")
print("="*80)

for row in range(50, 70):
    cell_a = ws[f'A{row}']
    print(f"Row {row}: A={repr(cell_a.value)[:60]}")
    
    # Show merges
    merges = [str(m) for m in ws.merged_cells.ranges if m.min_row == row]
    if merges:
        print(f"  Merges: {', '.join(merges)}")
