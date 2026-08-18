from openpyxl import load_workbook

template = load_workbook('static/reference/FSRFORMAT.xlsx')
ws_t = template.active

generated = load_workbook('generated_fsr/FSR_Devanadera_20260817_145821.xlsx')
ws_g = generated.active

print("\n=== DETAILED FONT COMPARISON ===\n")

cells_to_check = [
    ('A27', 'A20', "Concurrent teaching header"),
    ('A31', 'A24', "COLLEGE OUTSIDE U.P. SYSTEM"),
    ('A34', 'A27', "U.P. COLLEGE/DEPT."),
]

for t_cell, g_cell, desc in cells_to_check:
    print(f"\n{desc}:")
    print(f"  Template {t_cell}:")
    t = ws_t[t_cell]
    print(f"    Font name: {t.font.name}")
    print(f"    Font size: {t.font.size}")
    print(f"    Bold: {t.font.bold}")
    print(f"    Italic: {t.font.italic}")
    print(f"    Color: {t.font.color}")
    print(f"    Family: {t.font.family}")
    print(f"    Charset: {t.font.charset}")
    print(f"    Scheme: {t.font.scheme}")
    
    print(f"  Generated {g_cell}:")
    g = ws_g[g_cell]
    print(f"    Font name: {g.font.name}")
    print(f"    Font size: {g.font.size}")
    print(f"    Bold: {g.font.bold}")
    print(f"    Italic: {g.font.italic}")
    print(f"    Color: {g.font.color}")
    print(f"    Family: {g.font.family}")
    print(f"    Charset: {g.font.charset}")
    print(f"    Scheme: {g.font.scheme}")
    
    # Check if they match
    if (t.font.name != g.font.name or t.font.size != g.font.size or 
        t.font.bold != g.font.bold or str(t.font.scheme) != str(g.font.scheme)):
        print(f"  ⚠️ FONT PROPERTIES DIFFER!")
